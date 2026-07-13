"""
Csapat nyilvántartó Excel sablon generálása.
Létrehoz egy xlsx fájlt:
  - Legenda            - kódok, színek, használati útmutató
  - Napi jelenlét YYYY  - évenként külön munkalap, napi szintű szabadság/beteg/ünnep (16 fő)
  - Heti beosztás       - egy folytonos munkalap az összes évre, heti váltott műszak
                          (Éjjel / Délután / Délelőtt 3-hetes rotáció, mindenkire egyszerre)

Ez a fájl a "forrás" adat, amit a csoportvezető szerkeszt. A generate_html.py
script ebből készíti a csapat számára megtekinthető weboldalt.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

YEARS = [2025, 2026, 2027, 2028]
TODAY = datetime.date(2026, 7, 13)
EMPLOYEES = [
    "24009", "18639", "25856", "33501", "20161", "21319", "35883", "27556",
    "33373", "21180", "35885", "90008327", "35884", "29739", "22166",
]
N_EMPLOYEES = len(EMPLOYEES)
DEFAULT_MIN_STAFFING = 10
DEFAULT_VACATION_QUOTA = 20  # törvényi alap szabadságkeret napokban, soronként felülírható


def easter_sunday(year):
    """Húsvétvasárnap dátuma (Meeus/Jones/Butcher algoritmus, Gergely-naptár)."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime.date(year, month, day)


def hungarian_holidays(year):
    """Magyar hivatalos munkaszüneti napok listája egy adott évre."""
    easter = easter_sunday(year)
    return [
        datetime.date(year, 1, 1),                        # Újév
        datetime.date(year, 3, 15),                        # Nemzeti ünnep
        easter - datetime.timedelta(days=2),               # Nagypéntek
        easter + datetime.timedelta(days=1),                # Húsvéthétfő
        datetime.date(year, 5, 1),                         # A munka ünnepe
        easter + datetime.timedelta(days=50),               # Pünkösdhétfő
        datetime.date(year, 8, 20),                        # Állami ünnep
        datetime.date(year, 10, 23),                       # Nemzeti ünnep
        datetime.date(year, 11, 1),                        # Mindenszentek
        datetime.date(year, 12, 25),                       # Karácsony
        datetime.date(year, 12, 26),                       # Karácsony másnapja
    ]

DAILY_CODES = {
    "":     ("Munkanap", "FFFFFF"),
    "SZ":   ("Vállalati szabadság", "FFC000"),
    "SZMV": ("Saját szabadság", "FFD966"),
    "B":    ("Betegszabadság", "FF6B6B"),
    "KI":   ("Kivétel", "BDD7EE"),
    "UN":   ("Ünnepnap", "C39BD3"),
    "P":    ("Pihenőnap", "A9D18E"),
}
SUMMARY_CODES = [("SZ", "Vállalati szabadság"), ("SZMV", "Saját szabadság"), ("B", "Betegszabadság"), ("KI", "Kivétel"), ("UN", "Ünnepnap"), ("P", "Pihenőnap")]

# Váltott műszakrend: 3 hetes rotáció, az egész csapatra egyszerre érvényes.
# A TODAY-t tartalmazó hét = Éjjel, utána Délután, utána Délelőtt, majd újra Éjjel...
SHIFT_CYCLE = ["Éjjel", "Délután", "Délelőtt"]
WEEKLY_EXTRA_CODES = ["Vállalati szabadság", "Saját szabadság", "Pihenő", "Egyéb"]
WEEKLY_CODES = SHIFT_CYCLE + WEEKLY_EXTRA_CODES
WEEKLY_COLORS = {
    "Éjjel": "5B4B8A", "Délután": "F4B183", "Délelőtt": "FFE699",
    "Vállalati szabadság": "FFC000", "Saját szabadság": "FFD966", "Pihenő": "70AD47", "Egyéb": "D9D9D9",
}

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=9)
WEEKEND_FILL = PatternFill("solid", fgColor="E7E6E6")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF", size=10)
ALERT_FILL = PatternFill("solid", fgColor="FF6B6B")

HU_MONTHS = ["jan", "feb", "márc", "ápr", "máj", "jún", "júl", "aug", "szep", "okt", "nov", "dec"]

wb = Workbook()

# ---------------------------------------------------------------- Legenda
ws_legend = wb.active
ws_legend.title = "Legenda"
ws_legend.sheet_view.showGridLines = False
ws_legend["B2"] = "Csapat nyilvántartó — használati útmutató"
ws_legend["B2"].font = Font(name=FONT_NAME, bold=True, size=14)

ws_legend["B4"] = "1) Napi jelenlét (évenkénti lapok): minden nap / minden dolgozó cellájába írd be a kódot (legördülő listából is választható)."
ws_legend["B5"] = "2) Heti beosztás: a váltott műszakrend (Éjjel/Délután/Délelőtt) automatikusan ki van töltve 3 hetes rotációban, mindenkire egyszerre. Ahol valaki kivétel (szabin van — saját vagy vállalati —, pihenőn stb.), azt felülírhatod az adott cellában — ez automatikusan bekerül a weboldal 'Kivételek' listájába."
ws_legend["B6"] = "3) Éves összesítő: automatikusan számolja évenként/dolgozónként a vállalati szabadság, saját szabadság, betegszabadság, ünnepnap és pihenőnap napok számát. A sárga 'Szabadságkeret' cellák szerkeszthetők (alapértelmezett 20 nap) — a 'Hátralévő szabadság' ebből vonja le a felhasznált (vállalati + saját) szabadságnapokat."
ws_legend["B7"] = "4) A magyar munkaszüneti napok (Újév, Nagypéntek, Húsvét, Pünkösd, Aug.20, Okt.23, Nov.1, Karácsony stb.) minden évre automatikusan be vannak írva UN kóddal."
ws_legend["B8"] = "5) Minden napi jelenlét lap alján van egy 'Jelenlévők száma' sor és egy sárga 'Minimum létszám' cella — ha egy napon a jelenlévők száma a küszöb alá esik, az a cella pirosra vált. Hétvégén ez a sor nem releváns."
ws_legend["B9"] = "6) Mentsd el a fájlt, majd futtasd a generate_html.py scriptet — ez elkészíti a csapat számára a megtekinthető weboldalt (csapat_naptar.html)."
ws_legend["B10"] = "7) Az elkészült html fájlt oszd meg a csapattal (email, közös meghajtó, intranet) — ők csak megnézik, nem szerkesztik."
for r in range(4, 11):
    ws_legend[f"B{r}"].font = Font(name=FONT_NAME, size=10)
    ws_legend[f"B{r}"].alignment = Alignment(wrap_text=True)

ws_legend["B12"] = "Napi kódok:"
ws_legend["B12"].font = Font(name=FONT_NAME, bold=True, size=11)
row = 13
for code, (label, color) in DAILY_CODES.items():
    ws_legend[f"B{row}"] = code if code else "(üres)"
    ws_legend[f"C{row}"] = label
    ws_legend[f"B{row}"].fill = PatternFill("solid", fgColor=color)
    ws_legend[f"B{row}"].font = Font(name=FONT_NAME, size=10)
    ws_legend[f"C{row}"].font = Font(name=FONT_NAME, size=10)
    row += 1

row += 1
ws_legend[f"B{row}"] = "Heti műszakrend (3 hetes rotáció, automatikus):"
ws_legend[f"B{row}"].font = Font(name=FONT_NAME, bold=True, size=11)
row += 1
for shift in SHIFT_CYCLE:
    ws_legend[f"B{row}"] = shift
    ws_legend[f"B{row}"].fill = PatternFill("solid", fgColor=WEEKLY_COLORS[shift])
    ws_legend[f"B{row}"].font = Font(name=FONT_NAME, size=10)
    row += 1

row += 1
ws_legend[f"B{row}"] = "Kivétel esetén választható (felülírja a rotációt, bekerül a kivétel listába):"
ws_legend[f"B{row}"].font = Font(name=FONT_NAME, bold=True, size=11)
row += 1
for code in WEEKLY_EXTRA_CODES:
    ws_legend[f"B{row}"] = code
    ws_legend[f"B{row}"].fill = PatternFill("solid", fgColor=WEEKLY_COLORS[code])
    ws_legend[f"B{row}"].font = Font(name=FONT_NAME, size=10)
    row += 1

ws_legend.column_dimensions["A"].width = 2
ws_legend.column_dimensions["B"].width = 60
ws_legend.column_dimensions["C"].width = 22

# ---------------------------------------------------------------- Napi jelenlét (évenként)
for YEAR in YEARS:
    ws = wb.create_sheet(f"Napi jelenlét {YEAR}")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B2"

    start = datetime.date(YEAR, 1, 1)
    end = datetime.date(YEAR, 12, 31)
    n_days = (end - start).days + 1

    ws["A1"] = "Munkatárs"
    ws["A1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws.column_dimensions["A"].width = 16

    for d in range(n_days):
        day = start + datetime.timedelta(days=d)
        col = d + 2
        letter = get_column_letter(col)
        cell = ws.cell(row=1, column=col, value=f"{HU_MONTHS[day.month-1]}.{day.day}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(text_rotation=90, horizontal="center")
        ws.column_dimensions[letter].width = 3.2
        if day.weekday() >= 5:  # szombat, vasárnap
            for r in range(2, N_EMPLOYEES + 2):
                ws.cell(row=r, column=col).fill = WEEKEND_FILL

    for i, name in enumerate(EMPLOYEES):
        r = i + 2
        ws.cell(row=r, column=1, value=name).font = Font(name=FONT_NAME, size=10)

    data_range = f"B2:{get_column_letter(n_days + 1)}{N_EMPLOYEES + 1}"
    dv = DataValidation(type="list", formula1='"" ,SZ,SZMV,B,KI,UN,P', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(data_range)

    for code, (label, color) in DAILY_CODES.items():
        if not code:
            continue
        rule = CellIsRule(operator="equal", formula=[f'"{code}"'], fill=PatternFill("solid", fgColor=color))
        ws.conditional_formatting.add(data_range, rule)

    # magyar munkaszüneti napok automatikus beillesztése (UN)
    for holiday in hungarian_holidays(YEAR):
        col = 2 + (holiday - start).days
        for r in range(2, N_EMPLOYEES + 2):
            ws.cell(row=r, column=col, value="UN")

    # minta adat csak a 2026-os (jelenlegi) lapon, 2026.07 hónap körül
    if YEAR == 2026:
        sample_start_col = 2 + (datetime.date(YEAR, 7, 1) - start).days
        ws.cell(row=2, column=sample_start_col + 5, value="SZ")
        ws.cell(row=2, column=sample_start_col + 6, value="SZ")
        ws.cell(row=2, column=sample_start_col + 7, value="SZ")
        ws.cell(row=3, column=sample_start_col + 1, value="B")
        ws.cell(row=3, column=sample_start_col + 2, value="B")
        ws.cell(row=4, column=sample_start_col + 10, value="SZMV")
        ws.cell(row=4, column=sample_start_col + 11, value="SZMV")
        ws.cell(row=5, column=sample_start_col - 20, value="SZ")  # júniusi minta
        # minta alullétszámozott nap (2026.07.16): 7 fő szabin -> 9 fő marad, a 10-es küszöb alatt
        for emp_row in range(6, 13):
            ws.cell(row=emp_row, column=sample_start_col + 15, value="SZ")

    # napi jelenlévők száma (munkanapon dolgozók) + minimum létszám riasztás
    staff_row = N_EMPLOYEES + 2
    threshold_row = N_EMPLOYEES + 3
    last_col_letter = get_column_letter(n_days + 1)

    ws.cell(row=staff_row, column=1, value="Jelenlévők száma").font = Font(name=FONT_NAME, bold=True, size=9)
    ws.cell(row=threshold_row, column=1, value="Minimum létszám:").font = Font(name=FONT_NAME, bold=True, size=9)
    threshold_cell = ws.cell(row=threshold_row, column=2, value=DEFAULT_MIN_STAFFING)
    threshold_cell.fill = INPUT_FILL
    threshold_cell.font = INPUT_FONT
    threshold_ref = f"$B${threshold_row}"

    for d in range(n_days):
        col = d + 2
        letter = get_column_letter(col)
        c = ws.cell(row=staff_row, column=col, value=f"=COUNTBLANK({letter}2:{letter}{N_EMPLOYEES+1})")
        c.font = Font(name=FONT_NAME, size=9)
        c.alignment = Alignment(horizontal="center")

    staff_range = f"B{staff_row}:{last_col_letter}{staff_row}"
    ws.conditional_formatting.add(
        staff_range,
        FormulaRule(formula=[f"B{staff_row}<{threshold_ref}"], fill=ALERT_FILL, font=Font(name=FONT_NAME, color="FFFFFF", bold=True)),
    )

# ---------------------------------------------------------------- Heti beosztás (folytonos, minden évre)
ws2 = wb.create_sheet("Heti beosztás")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B2"
ws2["A1"] = "Munkatárs"
ws2["A1"].font = HEADER_FONT
ws2["A1"].fill = HEADER_FILL
ws2.column_dimensions["A"].width = 16

overall_start = datetime.date(YEARS[0], 1, 1)
overall_end = datetime.date(YEARS[-1], 12, 31)

# aktuális hét hétfője = a rotáció "nulladik" hete (Éjjel)
today_monday = TODAY - datetime.timedelta(days=TODAY.weekday())

week_starts = []
d = overall_start - datetime.timedelta(days=overall_start.weekday())
while d <= overall_end:
    week_starts.append(d)
    d += datetime.timedelta(days=7)

for w, wk in enumerate(week_starts):
    col = w + 2
    letter = get_column_letter(col)
    cell = ws2.cell(row=1, column=col, value=f"{wk.year}.{wk.month}.{wk.day}")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(text_rotation=90, horizontal="center")
    ws2.column_dimensions[letter].width = 4.5

for i, name in enumerate(EMPLOYEES):
    r = i + 2
    ws2.cell(row=r, column=1, value=name).font = Font(name=FONT_NAME, size=10)

dv2 = DataValidation(type="list", formula1='"' + ",".join(WEEKLY_CODES) + '"', allow_blank=True)
ws2.add_data_validation(dv2)
weekly_range = f"B2:{get_column_letter(len(week_starts)+1)}{N_EMPLOYEES+1}"
dv2.add(weekly_range)

for shift, color in WEEKLY_COLORS.items():
    rule = CellIsRule(operator="equal", formula=[f'"{shift}"'], fill=PatternFill("solid", fgColor=color))
    ws2.conditional_formatting.add(weekly_range, rule)

# automatikus rotáció kitöltése mindenkinek minden hétre
week_offset_to_shift = {}
for w, wk in enumerate(week_starts):
    offset = (wk - today_monday).days // 7
    shift = SHIFT_CYCLE[offset % 3]
    week_offset_to_shift[w] = shift

for i in range(N_EMPLOYEES):
    r = i + 2
    for w in range(len(week_starts)):
        ws2.cell(row=r, column=w + 2, value=week_offset_to_shift[w])

# minta kivételek: néhány felülírt hét, hogy a "Kivételek" lista lásson valamit
cur_week_col = week_starts.index(today_monday) + 2
ws2.cell(row=2, column=cur_week_col, value="Saját szabadság")        # 1. dolgozó: e héten szabin, nem Éjjel
ws2.cell(row=3, column=cur_week_col + 1, value="Pihenő")             # 2. dolgozó: jövő héten pihenő
ws2.cell(row=4, column=cur_week_col + 2, value="Vállalati szabadság") # 3. dolgozó: az azutáni héten vállalati szabin

# ---------------------------------------------------------------- Éves összesítő
ws3 = wb.create_sheet("Éves összesítő")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "B2"
ws3["A1"] = "Munkatárs"
ws3["A1"].font = HEADER_FONT
ws3["A1"].fill = HEADER_FILL
ws3.column_dimensions["A"].width = 16

col = 2
year_last_col = {}
year_block_start = {}
for YEAR in YEARS:
    year_block_start[YEAR] = col
    n_days = (datetime.date(YEAR, 12, 31) - datetime.date(YEAR, 1, 1)).days + 1
    year_last_col[YEAR] = get_column_letter(n_days + 1)
    headers = [label for _, label in SUMMARY_CODES] + ["Szabadságkeret", "Hátralévő szabadság"]
    for label in headers:
        letter = get_column_letter(col)
        cell = ws3.cell(row=1, column=col, value=f"{YEAR} {label}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(text_rotation=90, horizontal="center", wrap_text=True)
        ws3.column_dimensions[letter].width = 4.5
        col += 1

for i, name in enumerate(EMPLOYEES):
    r = i + 2
    ws3.cell(row=r, column=1, value=name).font = Font(name=FONT_NAME, size=10)
    for YEAR in YEARS:
        sheet_name = f"Napi jelenlét {YEAR}"
        last_col = year_last_col[YEAR]
        rng = f"'{sheet_name}'!B{r}:{last_col}{r}"
        col = year_block_start[YEAR]
        code_col_letter = {}  # kód -> oszlopbetű, hogy a hátralévő formula pontosan a SZ+SZMV oszlopokra hivatkozzon
        for code, label in SUMMARY_CODES:
            c = ws3.cell(row=r, column=col, value=f'=COUNTIF({rng},"{code}")')
            c.number_format = "0"
            c.font = Font(name=FONT_NAME, size=10)
            c.alignment = Alignment(horizontal="center")
            code_col_letter[code] = get_column_letter(col)
            col += 1
        # Szabadságkeret: szerkeszthető input, alapértelmezett 20 nap
        keret_cell = ws3.cell(row=r, column=col, value=DEFAULT_VACATION_QUOTA)
        keret_cell.number_format = "0"
        keret_cell.fill = INPUT_FILL
        keret_cell.font = INPUT_FONT
        keret_cell.alignment = Alignment(horizontal="center")
        keret_col_letter = get_column_letter(col)
        col += 1
        # Hátralévő szabadság = keret - (vállalati szabadság + saját szabadság) — a két típus együtt
        # terheli ugyanazt az éves keretet.
        hatra_cell = ws3.cell(
            row=r, column=col,
            value=f"={keret_col_letter}{r}-{code_col_letter['SZ']}{r}-{code_col_letter['SZMV']}{r}",
        )
        hatra_cell.number_format = "0"
        hatra_cell.font = Font(name=FONT_NAME, size=10, bold=True)
        hatra_cell.alignment = Alignment(horizontal="center")
        col += 1

wb.save("csapat_nyilvantartas.xlsx")
print("OK: csapat_nyilvantartas.xlsx elkészült")
print("Aktuális hét (", today_monday, ") műszakja:", week_offset_to_shift[week_starts.index(today_monday)])
