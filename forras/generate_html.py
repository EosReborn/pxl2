"""
Beolvassa a csapat_nyilvantartas.xlsx fájlt és elkészíti a csapat_naptar.html
weboldalt.

Ez a script csak az ALAP felépítéshez kell (új év hozzáadása, ünnepnapok
újraszámolása, teljes újraindítás). A napi haszálathoz NEM kell többé Excel
vagy ez a script: a kész weboldalon van egy "Szerkesztés" mód, ahol
közvetlenül a böngészőben lehet módosítani az adatokat. A "Mentés" gomb —
ha egyszer be van állítva egy GitHub personal access token — közvetlenül
visszaírja a módosítást a GitHub repóba (nincs letöltés/feltöltés). Beállítás
nélkül a Mentés gomb letölti az új, önálló index.html-t (kézi feltöltéshez).

Futtatás:  python3 generate_html.py
Bemenet:   csapat_nyilvantartas.xlsx  (a kiinduló adat)
Kimenet:   csapat_naptar.html        (ezt kell index.html néven feltölteni)
"""
import json
import os
import re
import datetime
from openpyxl import load_workbook

SRC = "csapat_nyilvantartas.xlsx"
OUT = "admin.html"  # teljes nézet — NEM index.html, csak akinek elküldöd a linket
INDEX_OUT = "index.html"  # nyilvános főoldal — nem tartalmaz személyes adatot
PEOPLE_DIR = "people"  # egy fájl / dolgozó, csak a saját adatával
TODAY = datetime.date(2026, 7, 13)

DAILY_CODES = {
    "":   {"label": "Munkanap", "color": "#ffffff"},
    "SZ": {"label": "Szabadság", "color": "#f5a623"},
    "B":  {"label": "Betegszabadság", "color": "#ef5350"},
    "OH": {"label": "Home office", "color": "#5b9bd5"},
    "UN": {"label": "Ünnepnap", "color": "#ab7fd1"},
    "P":  {"label": "Pihenőnap", "color": "#7cb872"},
}
SUMMARY_CODES = ["SZ", "B", "OH", "UN", "P"]
SHIFT_CYCLE = ["Éjjel", "Délután", "Délelőtt"]
SHIFT_COLORS = {
    "Éjjel": "#5b4b8a",
    "Délután": "#ef8f4e",
    "Délelőtt": "#f0c93b",
    "Szabadság": "#f5a623",
    "Home office": "#5b9bd5",
    "Pihenő": "#4d9a4d",
    "Egyéb": "#9aa1ab",
}
WEEKLY_CODES = SHIFT_CYCLE + ["Szabadság", "Home office", "Pihenő", "Egyéb"]

wb = load_workbook(SRC, data_only=True)

# --- napi jelenlét: minden "Napi jelenlét YYYY" lap beolvasása ---
years = []
daily_by_year = {}
for sheet_name in wb.sheetnames:
    m = re.match(r"Napi jelenlét (\d{4})", sheet_name)
    if not m:
        continue
    year = int(m.group(1))
    years.append(year)
    ws = wb[sheet_name]
    max_col = ws.max_column
    dates = [ws.cell(row=1, column=c).value for c in range(2, max_col + 1)]

    employees = []
    daily = {}
    r = 2
    while True:
        name = ws.cell(row=r, column=1).value
        if not name or name in ("Jelenlévők száma", "Minimum létszám:"):
            break
        employees.append(name)
        row_codes = []
        for c in range(2, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_codes.append((v or "").strip().upper())
        daily[name] = row_codes
        r += 1

    threshold_row = r + 1
    min_staffing = ws.cell(row=threshold_row, column=2).value or 0

    daily_by_year[year] = {"dates": dates, "employees": employees, "daily": daily, "min_staffing": min_staffing}
years.sort()

all_employees = daily_by_year[years[0]]["employees"]

# --- heti beosztás beolvasása ---
ws2 = wb["Heti beosztás"]
max_col2 = ws2.max_column
week_labels = []
week_years = []
for c in range(2, max_col2 + 1):
    label = ws2.cell(row=1, column=c).value
    week_labels.append(label)
    week_years.append(int(str(label).split(".")[0]))

weekly = {}
r = 2
while ws2.cell(row=r, column=1).value:
    name = ws2.cell(row=r, column=1).value
    row_vals = []
    for c in range(2, max_col2 + 1):
        v = ws2.cell(row=r, column=c).value
        row_vals.append(v or "")
    weekly[name] = row_vals
    r += 1

# --- szabadságkeret beolvasása (csak a keret-oszlopok, a felhasznált/hátralévő
#     mostantól élőben, a böngészőben számolódik az aktuális adatokból) ---
ws3 = wb["Éves összesítő"]
max_col3 = ws3.max_column
headers3 = [ws3.cell(row=1, column=c).value for c in range(2, max_col3 + 1)]
quota = {name: {} for name in all_employees}
r = 2
while ws3.cell(row=r, column=1).value:
    name = ws3.cell(row=r, column=1).value
    for idx, header in enumerate(headers3):
        y_str, label = str(header).split(" ", 1)
        if label == "Szabadságkeret":
            quota.setdefault(name, {})[int(y_str)] = ws3.cell(row=r, column=idx + 2).value or 0
    r += 1

today_huabbr = ["jan", "feb", "márc", "ápr", "máj", "jún", "júl", "aug", "szep", "okt", "nov", "dec"][TODAY.month - 1]
today_monday = TODAY - datetime.timedelta(days=TODAY.weekday())

data = {
    "generated": datetime.datetime.now().strftime("%Y.%m.%d %H:%M"),
    "years": years,
    "employees": all_employees,
    "daily_by_year": daily_by_year,
    "week_labels": week_labels,
    "week_years": week_years,
    "weekly": weekly,
    "codes": DAILY_CODES,
    "summary_codes": SUMMARY_CODES,
    "shift_colors": SHIFT_COLORS,
    "weekly_codes": WEEKLY_CODES,
    "quota": quota,
    "today_year": TODAY.year,
    "today_label": f"{today_huabbr}.{TODAY.day}",
    "today_week_label": f"{today_monday.year}.{today_monday.month}.{today_monday.day}",
    "today_display": f"{TODAY.year}. {['jan.','febr.','márc.','ápr.','máj.','jún.','júl.','aug.','szept.','okt.','nov.','dec.'][TODAY.month-1]} {TODAY.day}.",
}

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="hu">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Csapat naptár</title>
<style>
  :root {
    --border: #e5e7eb;
    --text: #1f2430;
    --muted: #6b7280;
    --accent: #4f46e5;
    --accent2: #7c3aed;
    --bg: #f4f5f9;
    --card: #ffffff;
    --radius: 14px;
    --shadow: 0 1px 3px rgba(17, 24, 39, 0.06), 0 1px 2px rgba(17, 24, 39, 0.08);
  }
  * { box-sizing: border-box; }
  body {
    font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    margin: 0;
    background: var(--bg);
    color: var(--text);
  }
  header {
    background: linear-gradient(120deg, var(--accent) 0%, var(--accent2) 100%);
    color: white;
    padding: 26px 32px 30px;
    display: flex; justify-content: space-between; align-items: flex-end; flex-wrap: wrap; gap: 14px;
  }
  header .titles h1 { margin: 0 0 6px 0; font-size: 23px; font-weight: 700; letter-spacing: -0.01em; }
  header .titles p { margin: 0; opacity: 0.88; font-size: 13px; }
  .header-actions { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
  .today-badge {
    background: rgba(255,255,255,0.16); border: 1px solid rgba(255,255,255,0.3);
    border-radius: 999px; padding: 7px 16px; font-size: 13px; font-weight: 600;
  }
  .hdr-btn {
    display: inline-flex; align-items: center; gap: 6px;
    background: rgba(255,255,255,0.16); border: 1px solid rgba(255,255,255,0.3); color: white;
    border-radius: 999px; padding: 7px 16px; font-size: 13px; font-weight: 700; cursor: pointer;
  }
  .hdr-btn:hover { background: rgba(255,255,255,0.28); }
  .hdr-btn.on { background: white; color: var(--accent); }
  .hdr-btn svg { width: 14px; height: 14px; }
  .wrap { max-width: 1440px; margin: -14px auto 0; padding: 0 22px 30px; }
  .edit-banner {
    display: none; background: #fff7e6; border: 1px solid #f5c26b; color: #8a5a00;
    border-radius: 10px; padding: 10px 16px; font-size: 13px; margin-bottom: 14px; font-weight: 600;
  }
  .edit-banner.show { display: block; }
  .tabs {
    display: flex; gap: 6px; margin-bottom: 16px; background: var(--card); border-radius: 999px;
    padding: 6px; box-shadow: var(--shadow); width: fit-content;
  }
  .tab-btn {
    display: flex; align-items: center; gap: 7px;
    padding: 9px 18px; border-radius: 999px; border: none;
    background: transparent; cursor: pointer; font-size: 13.5px; font-weight: 600;
    color: var(--muted); transition: all .15s ease;
  }
  .tab-btn:hover { background: #f1f1f6; }
  .tab-btn.active { background: linear-gradient(120deg, var(--accent), var(--accent2)); color: white; box-shadow: 0 2px 6px rgba(79,70,229,.35); }
  .tab-btn svg { width: 15px; height: 15px; flex-shrink: 0; }
  .panel { display: none; background: var(--card); border-radius: var(--radius); box-shadow: var(--shadow); padding: 20px; }
  .panel.active { display: block; }
  .controls { display: flex; gap: 14px; align-items: center; margin-bottom: 16px; flex-wrap: wrap; }
  .controls label { font-size: 13px; font-weight: 600; color: var(--muted); }
  select, input[type=text], input[type=number] {
    padding: 8px 12px; border-radius: 8px; border: 1px solid var(--border); font-size: 13.5px;
    background: #fafafd; outline: none;
  }
  select:focus, input:focus { border-color: var(--accent); background: white; }
  input[type=number] { width: 70px; }
  .btn-icon {
    display: inline-flex; align-items: center; gap: 6px; margin-left: auto;
    padding: 8px 14px; border-radius: 8px; border: 1px solid var(--border); background: white;
    cursor: pointer; font-size: 13px; font-weight: 600; color: var(--text);
  }
  .btn-icon:hover { background: #f4f4fb; border-color: var(--accent); }
  .btn-icon svg { width: 14px; height: 14px; }
  table { border-collapse: collapse; font-size: 12px; width: 100%; }
  th, td {
    border: 1px solid var(--border); padding: 4px 5px; text-align: center; white-space: nowrap;
  }
  th.name-col, td.name-col {
    position: sticky; left: 0; background: white; text-align: left; z-index: 2;
    min-width: 150px; font-weight: 600; font-size: 12.5px;
  }
  .avatar {
    display: inline-flex; align-items: center; justify-content: center;
    width: 22px; height: 22px; border-radius: 50%; background: #eef0fb; color: var(--accent);
    font-size: 10px; font-weight: 700; margin-right: 7px; vertical-align: middle;
  }
  thead th { position: sticky; top: 0; background: #383e73; color: white; z-index: 3; font-weight: 600; font-size: 11px; }
  thead th.name-col { z-index: 4; background: #383e73; }
  tbody tr:nth-child(even) td:not(.name-col):not([style]) { background: #fafafc; }
  tbody tr:hover td:not(.name-col) { filter: brightness(0.96); }
  tbody tr:hover td.name-col { background: #f4f4fb; }
  td.editable { cursor: pointer; }
  td.editable:hover { outline: 2px solid var(--accent); outline-offset: -2px; }
  td.editable-name { cursor: pointer; }
  td.editable-name:hover { background: #eef0fb; text-decoration: underline; text-decoration-style: dotted; }
  .editable-name-inline { cursor: pointer; }
  .editable-name-inline:hover { text-decoration: underline; text-decoration-style: dotted; }
  .table-scroll { overflow: auto; max-height: 65vh; border: 1px solid var(--border); border-radius: 10px; }
  .staff-row td { font-weight: 700; background: #eef0f7; }
  .staff-row td.alert { background: #ef5350 !important; color: white; }
  .legend { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 16px; }
  .legend span.chip {
    display: inline-flex; align-items: center; gap: 6px; font-size: 12.5px; padding: 5px 12px;
    border-radius: 999px; background: #f4f4fb; border: 1px solid var(--border);
  }
  .legend span.swatch { display: inline-block; width: 11px; height: 11px; border-radius: 50%; border: 1px solid rgba(0,0,0,0.1); }
  footer { text-align: center; color: var(--muted); font-size: 12px; padding: 20px 24px 30px; }
  .year-bar { display: flex; justify-content: center; gap: 8px; margin: 26px auto 4px; flex-wrap: wrap; }
  .year-btn {
    padding: 9px 22px; border-radius: 999px; border: 1px solid var(--border);
    background: white; cursor: pointer; font-size: 13.5px; font-weight: 700; color: var(--muted);
    box-shadow: var(--shadow); transition: all .15s ease;
  }
  .year-btn:hover { color: var(--accent); }
  .year-btn.active { background: linear-gradient(120deg, var(--accent), var(--accent2)); color: white; border-color: transparent; }
  .summary-cards { display: flex; flex-wrap: wrap; gap: 14px; }
  .summary-card {
    flex: 1 1 210px; border: 1px solid var(--border); border-radius: 12px; padding: 16px 18px;
    background: linear-gradient(180deg, #fafbff 0%, #ffffff 100%);
  }
  .summary-card h3 { margin: 0 0 12px 0; font-size: 14px; display: flex; align-items: center; }
  .summary-card table { width: 100%; font-size: 12.5px; }
  .summary-card td { border: none; padding: 3px 0; text-align: left; }
  .summary-card td.num { text-align: right; font-weight: 700; }
  .summary-card .quota-box {
    margin-top: 10px; padding-top: 10px; border-top: 1px dashed var(--border);
    display: flex; justify-content: space-between; align-items: center;
  }
  .summary-card .quota-remaining { font-size: 20px; font-weight: 800; color: var(--accent); }
  .summary-card .quota-caption { font-size: 11px; color: var(--muted); }
  .empty-note { color: var(--muted); font-size: 13px; padding: 10px 0 16px; }
  td.today-col { box-shadow: inset 0 0 0 2px var(--accent); font-weight: 700; }
  th.today-col { background: #2f2b6b !important; }
  .picker {
    position: absolute; z-index: 50; background: white; border-radius: 10px; box-shadow: 0 8px 24px rgba(0,0,0,.18);
    border: 1px solid var(--border); padding: 6px; display: flex; flex-direction: column; gap: 2px; min-width: 150px;
  }
  .picker button {
    display: flex; align-items: center; gap: 8px; border: none; background: transparent; text-align: left;
    padding: 7px 10px; border-radius: 6px; cursor: pointer; font-size: 12.5px; font-weight: 600; color: var(--text);
  }
  .picker button:hover { background: #f4f4fb; }
  .picker .dot { width: 10px; height: 10px; border-radius: 50%; border: 1px solid rgba(0,0,0,.1); flex-shrink: 0; }
  .modal-overlay {
    display: none; position: fixed; inset: 0; background: rgba(20,20,30,.5); z-index: 100;
    align-items: center; justify-content: center;
  }
  .modal-overlay.show { display: flex; }
  .modal {
    background: white; border-radius: 14px; padding: 24px 26px; width: 380px; max-width: 92vw;
    box-shadow: 0 20px 50px rgba(0,0,0,.25);
  }
  .modal h3 { margin: 0 0 6px 0; font-size: 16px; }
  .modal p.hint { font-size: 12px; color: var(--muted); margin: 0 0 16px 0; line-height: 1.5; }
  .modal label { display: block; font-size: 12.5px; font-weight: 600; color: var(--muted); margin: 10px 0 4px; }
  .modal input { width: 100%; padding: 9px 11px; border-radius: 8px; border: 1px solid var(--border); font-size: 13.5px; }
  .modal .modal-actions { display: flex; justify-content: space-between; align-items: center; margin-top: 20px; gap: 10px; }
  .modal .modal-actions .right { display: flex; gap: 8px; }
  .modal button.primary {
    background: linear-gradient(120deg, var(--accent), var(--accent2)); color: white; border: none;
    padding: 9px 18px; border-radius: 8px; font-weight: 700; cursor: pointer; font-size: 13px;
  }
  .modal button.secondary {
    background: #f1f1f6; color: var(--text); border: none; padding: 9px 16px; border-radius: 8px;
    font-weight: 600; cursor: pointer; font-size: 13px;
  }
  .modal button.link { background: none; border: none; color: #b3261e; font-size: 12px; cursor: pointer; padding: 4px 0; text-decoration: underline; }
  .toast {
    position: fixed; bottom: 24px; left: 50%; transform: translateX(-50%); z-index: 200;
    background: #1f2430; color: white; padding: 12px 20px; border-radius: 10px; font-size: 13.5px;
    box-shadow: 0 8px 24px rgba(0,0,0,.3); max-width: 90vw; text-align: center;
  }
  @media print {
    .header-actions, .tabs, .controls, .legend, .year-bar, footer, .btn-icon, .edit-banner { display: none !important; }
    body { background: white; }
    .wrap { margin: 0; padding: 0; max-width: 100%; }
    .panel { display: none !important; box-shadow: none; border-radius: 0; padding: 0; }
    .panel.active { display: block !important; }
    .table-scroll { max-height: none; overflow: visible; border: none; }
    header { background: white !important; color: black; padding: 10px 0; }
    thead th { background: #383e73 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  }
</style>
</head>
<body>
<script type="application/json" id="app-data">__DATA_JSON__</script>
<header>
  <div class="titles">
    <h1>Csapat naptár — szabadság, betegség, munkarend</h1>
    <p>Utolsó frissítés: __GENERATED__ &nbsp;·&nbsp; <span id="modeLabel">csak megtekintésre</span></p>
  </div>
  <div class="header-actions">
    <div class="today-badge">Ma: __TODAY_DISPLAY__</div>
    <button class="hdr-btn" id="editToggleBtn">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M17 3a2.85 2.83 0 1 1 4 4L7.5 20.5 2 22l1.5-5.5Z"/></svg>
      Szerkesztés
    </button>
    <button class="hdr-btn" id="saveBtn">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2Z"/><path d="M17 21v-8H7v8M7 3v5h8"/></svg>
      Mentés
    </button>
    <button class="hdr-btn" id="settingsBtn" title="GitHub mentés beállítása">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 1 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 1 1-2.83-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 1 1 2.83-2.83l.06.06A1.65 1.65 0 0 0 9 4.6a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 1 1 2.83 2.83l-.06.06A1.65 1.65 0 0 0 19.4 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1Z"/></svg>
    </button>
  </div>
</header>
<div class="wrap">
  <div class="edit-banner" id="editBanner">Szerkesztés mód bekapcsolva — kattints egy cellára a módosításhoz. A "Mentés" gomb közvetlenül visszaírja a változást a GitHub repóba.</div>

  <div class="tabs">
    <button class="tab-btn active" data-tab="daily">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><rect x="3" y="5" width="18" height="16" rx="2"/><path d="M3 10h18M8 3v4M16 3v4"/></svg>
      Napi jelenlét
    </button>
    <button class="tab-btn" data-tab="weekly">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 3"/></svg>
      Heti beosztás
    </button>
    <button class="tab-btn" data-tab="exceptions">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.3 3.9 1.8 18a2 2 0 0 0 1.7 3h17a2 2 0 0 0 1.7-3L13.7 3.9a2 2 0 0 0-3.4 0Z"/><path d="M12 9v4M12 17h.01"/></svg>
      Kivételek
    </button>
    <button class="tab-btn" data-tab="summary">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><path d="M3 3v18h18M8 17V10M13 17V6M18 17v-4"/></svg>
      Éves összesítő
    </button>
  </div>

  <div id="daily" class="panel active">
    <div class="controls">
      <label>Hónap</label>
      <select id="monthSelect"></select>
      <label>Keresés</label>
      <input type="text" id="searchDaily" placeholder="Munkatárs neve...">
      <label id="minStaffLabel" style="display:none;">Min. létszám</label>
      <input type="number" id="minStaffInput" style="display:none;">
      <button class="btn-icon" onclick="window.print()">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M6 9V2h12v7M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2M6 14h12v8H6z"/></svg>
        Nyomtatás / PDF
      </button>
    </div>
    <div class="table-scroll"><table id="dailyTable"></table></div>
    <div class="legend" id="legend"></div>
  </div>

  <div id="weekly" class="panel">
    <div class="controls">
      <label>Keresés</label>
      <input type="text" id="searchWeekly" placeholder="Munkatárs neve...">
      <button class="btn-icon" onclick="window.print()">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M6 9V2h12v7M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2M6 14h12v8H6z"/></svg>
        Nyomtatás / PDF
      </button>
    </div>
    <div class="table-scroll"><table id="weeklyTable"></table></div>
    <div class="legend" id="legendWeekly"></div>
  </div>

  <div id="exceptions" class="panel">
    <div class="controls">
      <label>Keresés</label>
      <input type="text" id="searchExc" placeholder="Munkatárs neve...">
    </div>
    <p class="empty-note">Azok a hetek, ahol a tényleges beosztás eltér az automatikus Éjjel→Délután→Délelőtt rotációtól (pl. szabadság, pihenő, home office, egyéb).</p>
    <div class="table-scroll"><table id="excTable"></table></div>
  </div>

  <div id="summary" class="panel">
    <div class="summary-cards" id="summaryCards"></div>
  </div>

  <div class="year-bar" id="yearBar"></div>
</div>
<footer>A jobb felső "Szerkesztés" móddal közvetlenül itt módosíthatod az adatokat, a "Mentés" pedig egy kattintással visszaírja a GitHub repóba.</footer>

<div class="modal-overlay" id="settingsOverlay">
  <div class="modal">
    <h3>GitHub mentés beállítása</h3>
    <p class="hint">Ezt csak egyszer kell beállítanod (ezen a gépen, ebben a böngészőben marad meg). A tokent a GitHub-on hozod létre: Settings → Developer settings → Fine-grained tokens → csak erre a repóra, "Contents: Read and write" jogosultsággal.</p>
    <label>GitHub felhasználó / szervezet</label>
    <input id="cfgOwner" placeholder="pl. kochnorbert">
    <label>Repó neve</label>
    <input id="cfgRepo" placeholder="pl. csapat-naptar">
    <label>Branch</label>
    <input id="cfgBranch" placeholder="main" value="main">
    <label>Fájl elérési útja a repóban</label>
    <input id="cfgPath" placeholder="admin.html" value="admin.html">
    <label>Personal access token</label>
    <input id="cfgToken" type="password" placeholder="ghp_...">
    <div class="modal-actions">
      <button class="link" id="cfgClear">Kijelentkezés / adatok törlése</button>
      <div class="right">
        <button class="secondary" id="cfgCancel">Mégse</button>
        <button class="primary" id="cfgSave">Mentés</button>
      </div>
    </div>
  </div>
</div>

<script>
const DATA = JSON.parse(document.getElementById('app-data').textContent);
let currentYear = DATA.years.includes(DATA.today_year) ? DATA.today_year : DATA.years[0];
let editMode = false;
const PERSONAL_TEMPLATE = __PERSONAL_TEMPLATE_JSON__;

document.querySelectorAll('.tab-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById(btn.dataset.tab).classList.add('active');
    closePicker();
  });
});

// --- GitHub mentési beállítások (csak ebben a böngészőben tárolva) ---
const CFG_KEY = 'csapatNaptarGithubCfg';
function loadCfg() { try { return JSON.parse(localStorage.getItem(CFG_KEY) || 'null'); } catch (e) { return null; } }
function saveCfg(cfg) { localStorage.setItem(CFG_KEY, JSON.stringify(cfg)); }
function clearCfg() { localStorage.removeItem(CFG_KEY); }
function hasCfg() { const c = loadCfg(); return !!(c && c.owner && c.repo && c.token); }

function showToast(msg, ms) {
  const t = document.createElement('div');
  t.className = 'toast';
  t.textContent = msg;
  document.body.appendChild(t);
  setTimeout(() => t.remove(), ms || 4500);
}

let pendingEnableEdit = false;
function openSettings(enableEditAfter) {
  pendingEnableEdit = !!enableEditAfter;
  const c = loadCfg() || {};
  document.getElementById('cfgOwner').value = c.owner || '';
  document.getElementById('cfgRepo').value = c.repo || '';
  document.getElementById('cfgBranch').value = c.branch || 'main';
  document.getElementById('cfgPath').value = c.path || 'admin.html';
  document.getElementById('cfgToken').value = c.token || '';
  document.getElementById('settingsOverlay').classList.add('show');
}
function closeSettings() { document.getElementById('settingsOverlay').classList.remove('show'); }

document.getElementById('settingsBtn').addEventListener('click', () => openSettings(false));
document.getElementById('cfgCancel').addEventListener('click', closeSettings);
document.getElementById('settingsOverlay').addEventListener('click', (e) => { if (e.target.id === 'settingsOverlay') closeSettings(); });
document.getElementById('cfgClear').addEventListener('click', () => {
  clearCfg();
  editMode = false;
  updateEditUI();
  closeSettings();
  showToast('Törölve — a beállítások és a token eltávolítva ebből a böngészőből.');
});
document.getElementById('cfgSave').addEventListener('click', () => {
  const cfg = {
    owner: document.getElementById('cfgOwner').value.trim(),
    repo: document.getElementById('cfgRepo').value.trim(),
    branch: document.getElementById('cfgBranch').value.trim() || 'main',
    path: document.getElementById('cfgPath').value.trim() || 'admin.html',
    token: document.getElementById('cfgToken').value.trim(),
  };
  if (!cfg.owner || !cfg.repo || !cfg.token) { alert('Töltsd ki a felhasználót, a repó nevét és a tokent.'); return; }
  if (cfg.path === 'index.html') {
    const ok = confirm('Figyelem: az "index.html" a publikus, adatmentes főoldal! Ha erre mented a teljes csapatnaptárt, mindenki adata nyilvánossá válik. Biztosan ezt akarod (nem inkább "admin.html")?');
    if (!ok) return;
  }
  saveCfg(cfg);
  closeSettings();
  if (pendingEnableEdit) { editMode = true; updateEditUI(); }
  showToast('Beállítás mentve ebben a böngészőben.');
});

// --- edit mode toggle ---
function updateEditUI() {
  document.getElementById('editToggleBtn').classList.toggle('on', editMode);
  document.getElementById('editBanner').classList.toggle('show', editMode);
  document.getElementById('modeLabel').textContent = editMode ? 'szerkesztés mód' : 'csak megtekintésre';
  document.getElementById('minStaffLabel').style.display = editMode ? '' : 'none';
  document.getElementById('minStaffInput').style.display = editMode ? '' : 'none';
  renderAll();
}
document.getElementById('editToggleBtn').addEventListener('click', () => {
  if (!editMode) {
    if (!hasCfg()) { openSettings(true); return; }
    editMode = true;
  } else {
    editMode = false;
  }
  updateEditUI();
});

// --- mentés: ha van GitHub beállítás, közvetlenül a repóba írja; ha nincs, letölti a fájlt ---
function currentHtml() {
  document.getElementById('app-data').textContent = JSON.stringify(DATA);
  return '<!DOCTYPE html>\\n' + document.documentElement.outerHTML;
}
function downloadFile(html) {
  const blob = new Blob([html], { type: 'text/html' });
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'index.html';
  document.body.appendChild(a);
  a.click();
  a.remove();
}
function slugify(name) {
  const repl = { 'á':'a','é':'e','í':'i','ó':'o','ö':'o','ő':'o','ú':'u','ü':'u','ű':'u' };
  let s = name.toLowerCase();
  Object.entries(repl).forEach(([k, v]) => { s = s.split(k).join(v); });
  s = s.replace(/[^a-z0-9]+/g, '-').replace(/^-+|-+$/g, '');
  return s;
}
function replaceAllStr(str, search, val) { return str.split(search).join(val); }

function buildPersonalHtml(name) {
  const cssBlock = document.querySelector('style').textContent;
  const pDaily = {};
  DATA.years.forEach(y => {
    const yd = DATA.daily_by_year[y];
    pDaily[y] = { dates: yd.dates, daily: { [name]: yd.daily[name] } };
  });
  const pData = {
    generated: DATA.generated || new Date().toLocaleString('hu-HU'),
    years: DATA.years,
    employee: name,
    daily_by_year: pDaily,
    week_labels: DATA.week_labels,
    week_years: DATA.week_years,
    weekly: { [name]: DATA.weekly[name] || [] },
    codes: DATA.codes,
    summary_codes: DATA.summary_codes,
    shift_colors: DATA.shift_colors,
    quota: { [name]: (DATA.quota[name] || {}) },
    today_year: DATA.today_year,
    today_label: DATA.today_label,
    today_week_label: DATA.today_week_label,
    today_display: DATA.today_display,
  };
  let out = PERSONAL_TEMPLATE;
  out = replaceAllStr(out, '__P_CSS_BLOCK__', cssBlock);
  out = replaceAllStr(out, '__P_NAME__', name);
  out = replaceAllStr(out, '__P_GENERATED__', pData.generated);
  out = replaceAllStr(out, '__P_TODAY_DISPLAY__', DATA.today_display);
  out = replaceAllStr(out, '__P_DATA_JSON__', JSON.stringify(pData));
  return out;
}

// --- generikus GET-sha + PUT segédfüggvény (admin.html és people/*.html mentéséhez is) ---
async function putFileToGithub(cfg, path, content, message) {
  const apiBase = `https://api.github.com/repos/${cfg.owner}/${cfg.repo}/contents/${path}`;
  const headers = { Authorization: `Bearer ${cfg.token}`, Accept: 'application/vnd.github+json' };
  const getResp = await fetch(`${apiBase}?ref=${encodeURIComponent(cfg.branch)}&t=${Date.now()}`, { headers, cache: 'no-store' });
  let sha;
  if (getResp.ok) {
    sha = (await getResp.json()).sha;
  } else if (getResp.status !== 404) {
    throw new Error(`Nem sikerült lekérni a(z) ${path} fájlt (${getResp.status}).`);
  }
  const contentB64 = btoa(unescape(encodeURIComponent(content)));
  const body = { message, content: contentB64, branch: cfg.branch };
  if (sha) body.sha = sha;
  const putResp = await fetch(apiBase, {
    method: 'PUT',
    headers: { ...headers, 'Content-Type': 'application/json' },
    body: JSON.stringify(body),
  });
  if (!putResp.ok) {
    const errJson = await putResp.json().catch(() => ({}));
    if (putResp.status === 409) {
      throw new Error(`A(z) ${path} időközben megváltozott a GitHub-on. Frissítsd az oldalt (Cmd/Ctrl+Shift+R), és próbáld újra.`);
    }
    throw new Error(`Mentés sikertelen (${path}, ${putResp.status}): ${errJson.message || 'ismeretlen hiba'}`);
  }
}

async function syncPersonalPages(cfg) {
  const results = { ok: 0, fail: [] };
  for (const name of DATA.employees) {
    const path = `people/${slugify(name)}.html`;
    try {
      await putFileToGithub(cfg, path, buildPersonalHtml(name), `Személyes oldal frissítve: ${name}`);
      results.ok++;
    } catch (err) {
      results.fail.push(name + ': ' + err.message);
    }
  }
  return results;
}

async function publishToGithub(cfg) {
  const html = currentHtml();
  showToast('Mentés a GitHub-ra...', 60000);
  try {
    await putFileToGithub(cfg, cfg.path, html, 'Csapat naptár frissítve – ' + new Date().toLocaleString('hu-HU'));
    document.querySelectorAll('.toast').forEach(t => t.remove());
    showToast(`Admin mentve, személyes oldalak szinkronizálása (${DATA.employees.length} db)...`, 60000);
    const results = await syncPersonalPages(cfg);
    document.querySelectorAll('.toast').forEach(t => t.remove());
    if (results.fail.length === 0) {
      showToast(`Mentve! Az admin és mind a ${results.ok} személyes oldal frissült. Az élő oldal 1-2 percen belül tükrözi.`);
    } else {
      alert(`Az admin mentve, de ${results.fail.length} személyes oldal szinkronizálása nem sikerült:\\n` + results.fail.join('\\n'));
    }
  } catch (err) {
    document.querySelectorAll('.toast').forEach(t => t.remove());
    alert('Hiba történt a GitHub mentés közben:\\n' + err.message + '\\n\\nLetöltöm helyette a fájlt, hogy kézzel is fel tudd tölteni.');
    downloadFile(html);
  }
}
document.getElementById('saveBtn').addEventListener('click', () => {
  const cfg = loadCfg();
  if (cfg && cfg.token) {
    publishToGithub(cfg);
  } else {
    downloadFile(currentHtml());
  }
});

// --- year bar ---
const yearBar = document.getElementById('yearBar');
DATA.years.forEach(y => {
  const btn = document.createElement('button');
  btn.className = 'year-btn' + (y === currentYear ? ' active' : '');
  btn.textContent = y;
  btn.dataset.year = y;
  btn.addEventListener('click', () => {
    currentYear = y;
    document.querySelectorAll('.year-btn').forEach(b => b.classList.toggle('active', parseInt(b.dataset.year) === y));
    renderAll();
  });
  yearBar.appendChild(btn);
});

// --- legend ---
const legendEl = document.getElementById('legend');
Object.entries(DATA.codes).forEach(([code, info]) => {
  if (!code) return;
  const el = document.createElement('span');
  el.className = 'chip';
  el.innerHTML = `<span class="swatch" style="background:${info.color}"></span>${info.label} (${code})`;
  legendEl.appendChild(el);
});
const legendWeeklyEl = document.getElementById('legendWeekly');
Object.entries(DATA.shift_colors).forEach(([label, color]) => {
  const el = document.createElement('span');
  el.className = 'chip';
  el.innerHTML = `<span class="swatch" style="background:${color}"></span>${label}`;
  legendWeeklyEl.appendChild(el);
});

// --- month select ---
const monthNames = ['Január','Február','Március','Április','Május','Június','Július','Augusztus','Szeptember','Október','November','December'];
const huMonthAbbr = ['jan','feb','márc','ápr','máj','jún','júl','aug','szep','okt','nov','dec'];
const monthSelect = document.getElementById('monthSelect');
monthNames.forEach((name, idx) => {
  const opt = document.createElement('option');
  opt.value = idx; opt.textContent = name;
  monthSelect.appendChild(opt);
});
monthSelect.value = huMonthAbbr.indexOf(DATA.today_label.split('.')[0]);

function monthOfIndex(dates, i) { return huMonthAbbr.indexOf(dates[i].split('.')[0]); }
function initials(name) {
  const parts = name.replace('Munkatárs ', '').trim().split(/\s+/).filter(Boolean);
  if (parts.length === 0) return '?';
  if (parts.length === 1) return parts[0].slice(0, 2).toUpperCase();
  return (parts[0][0] + parts[parts.length - 1][0]).toUpperCase();
}

// --- munkatárs átnevezése (minden adatstruktúrában konzisztensen) ---
function renameEmployee(oldName) {
  const input = prompt('Új név:', oldName);
  if (input === null) return;
  const newName = input.trim();
  if (!newName || newName === oldName) return;
  if (DATA.employees.includes(newName)) { alert('Már létezik ilyen nevű munkatárs.'); return; }
  DATA.employees = DATA.employees.map(n => n === oldName ? newName : n);
  DATA.years.forEach(y => {
    const yd = DATA.daily_by_year[y];
    yd.employees = yd.employees.map(n => n === oldName ? newName : n);
    if (yd.daily[oldName] !== undefined) { yd.daily[newName] = yd.daily[oldName]; delete yd.daily[oldName]; }
  });
  if (DATA.weekly[oldName] !== undefined) { DATA.weekly[newName] = DATA.weekly[oldName]; delete DATA.weekly[oldName]; }
  if (DATA.quota[oldName] !== undefined) { DATA.quota[newName] = DATA.quota[oldName]; delete DATA.quota[oldName]; }
  renderAll();
}

// --- élő számítások (mindig az aktuális DATA-ból, hogy szerkesztés után is pontos legyen) ---
function expectedShift(weekLabel) {
  const [y, mo, da] = weekLabel.split('.').map(Number);
  const [ty, tmo, tda] = DATA.today_week_label.split('.').map(Number);
  const days = Math.round((Date.UTC(y, mo - 1, da) - Date.UTC(ty, tmo - 1, tda)) / 86400000);
  const offset = Math.round(days / 7);
  const idx = ((offset % 3) + 3) % 3;
  return ['Éjjel', 'Délután', 'Délelőtt'][idx];
}
function computeStaffCounts(year) {
  const yd = DATA.daily_by_year[year];
  return yd.dates.map((d, i) => yd.employees.filter(name => !(yd.daily[name][i] || '')).length);
}
function computeSummary(name, year) {
  const yd = DATA.daily_by_year[year];
  const counts = { SZ: 0, B: 0, OH: 0, UN: 0, P: 0 };
  const codes = (yd.daily[name] || []);
  codes.forEach(c => { if (counts[c] !== undefined) counts[c]++; });
  return counts;
}

// --- popover a cellák szerkesztéséhez ---
let pickerEl = null;
function closePicker() { if (pickerEl) { pickerEl.remove(); pickerEl = null; } }
document.addEventListener('click', (e) => {
  if (pickerEl && !pickerEl.contains(e.target) && !e.target.classList.contains('editable')) closePicker();
});
function openPicker(cellEl, options, onSelect) {
  closePicker();
  pickerEl = document.createElement('div');
  pickerEl.className = 'picker';
  options.forEach(opt => {
    const b = document.createElement('button');
    b.innerHTML = `<span class="dot" style="background:${opt.color || '#fff'}"></span>${opt.label}`;
    b.addEventListener('click', (ev) => { ev.stopPropagation(); onSelect(opt.value); closePicker(); });
    pickerEl.appendChild(b);
  });
  document.body.appendChild(pickerEl);
  const rect = cellEl.getBoundingClientRect();
  pickerEl.style.left = (rect.left + window.scrollX) + 'px';
  pickerEl.style.top = (rect.bottom + window.scrollY + 4) + 'px';
}

// --- napi jelenlét ---
function renderDaily() {
  const yearData = DATA.daily_by_year[currentYear];
  if (!yearData) return;
  const monthIdx = parseInt(monthSelect.value, 10);
  const search = document.getElementById('searchDaily').value.trim().toLowerCase();
  const colIndexes = [];
  yearData.dates.forEach((d, i) => { if (monthOfIndex(yearData.dates, i) === monthIdx) colIndexes.push(i); });
  const isToday = (i) => currentYear === DATA.today_year && yearData.dates[i] === DATA.today_label;

  document.getElementById('minStaffInput').value = yearData.min_staffing;

  let html = '<thead><tr><th class="name-col">Munkatárs</th>';
  colIndexes.forEach(i => {
    html += `<th class="${isToday(i) ? 'today-col' : ''}">${yearData.dates[i].split('.')[1]}</th>`;
  });
  html += '</tr></thead><tbody>';

  yearData.employees.forEach(name => {
    if (search && !name.toLowerCase().includes(search)) return;
    const nameCls = editMode ? 'name-col editable-name' : 'name-col';
    html += `<tr><td class="${nameCls}" data-name="${name}" title="${editMode ? 'Kattints az átnevezéshez' : ''}"><span class="avatar">${initials(name)}</span>${name}</td>`;
    colIndexes.forEach(i => {
      const code = (yearData.daily[name] && yearData.daily[name][i]) || '';
      const info = DATA.codes[code];
      const bg = (info && code) ? `background:${info.color}${code==='B'?';color:white':''};` : '';
      const cls = (isToday(i) ? 'today-col ' : '') + (editMode ? 'editable' : '');
      html += `<td class="${cls}" data-emp="${name}" data-idx="${i}" style="${bg}" title="${info ? info.label : ''}">${code}</td>`;
    });
    html += '</tr>';
  });

  const staffCounts = computeStaffCounts(currentYear);
  html += `<tr class="staff-row"><td class="name-col">Jelenlévők száma</td>`;
  colIndexes.forEach(i => {
    const count = staffCounts[i];
    const low = count < yearData.min_staffing;
    html += `<td class="${low ? 'alert' : ''}" title="Minimum létszám: ${yearData.min_staffing}">${count}</td>`;
  });
  html += '</tr></tbody>';
  document.getElementById('dailyTable').innerHTML = html;

  if (editMode) {
    document.querySelectorAll('#dailyTable td.editable').forEach(cell => {
      cell.addEventListener('click', (e) => {
        e.stopPropagation();
        const name = cell.dataset.emp, idx = parseInt(cell.dataset.idx, 10);
        const options = Object.entries(DATA.codes).map(([code, info]) => ({ value: code, label: info.label + (code ? ` (${code})` : ''), color: info.color }));
        openPicker(cell, options, (val) => {
          DATA.daily_by_year[currentYear].daily[name][idx] = val;
          renderAll();
        });
      });
    });
    document.querySelectorAll('#dailyTable td.editable-name').forEach(cell => {
      cell.addEventListener('click', (e) => { e.stopPropagation(); renameEmployee(cell.dataset.name); });
    });
  }
}

// --- heti beosztás ---
function renderWeekly() {
  const search = document.getElementById('searchWeekly').value.trim().toLowerCase();
  const colIndexes = [];
  DATA.week_labels.forEach((w, i) => { if (DATA.week_years[i] === currentYear) colIndexes.push(i); });

  let html = '<thead><tr><th class="name-col">Munkatárs</th>';
  colIndexes.forEach(i => {
    const parts = DATA.week_labels[i].split('.');
    const isToday = DATA.week_labels[i] === DATA.today_week_label;
    html += `<th class="${isToday ? 'today-col' : ''}">${parts[1]}.${parts[2]}</th>`;
  });
  html += '</tr></thead><tbody>';
  DATA.employees.forEach(name => {
    if (search && !name.toLowerCase().includes(search)) return;
    const vals = DATA.weekly[name] || [];
    const nameCls = editMode ? 'name-col editable-name' : 'name-col';
    html += `<tr><td class="${nameCls}" data-name="${name}" title="${editMode ? 'Kattints az átnevezéshez' : ''}"><span class="avatar">${initials(name)}</span>${name}</td>`;
    colIndexes.forEach(i => {
      const v = vals[i] || '';
      const color = DATA.shift_colors[v];
      const isToday = DATA.week_labels[i] === DATA.today_week_label;
      const style = color ? `background:${color}${v === 'Éjjel' ? ';color:white' : ''};` : '';
      const cls = (isToday ? 'today-col ' : '') + (editMode ? 'editable' : '');
      html += `<td class="${cls}" data-emp="${name}" data-idx="${i}" style="${style}">${v}</td>`;
    });
    html += '</tr>';
  });
  html += '</tbody>';
  document.getElementById('weeklyTable').innerHTML = html;

  if (editMode) {
    document.querySelectorAll('#weeklyTable td.editable-name').forEach(cell => {
      cell.addEventListener('click', (e) => { e.stopPropagation(); renameEmployee(cell.dataset.name); });
    });
    document.querySelectorAll('#weeklyTable td.editable').forEach(cell => {
      cell.addEventListener('click', (e) => {
        e.stopPropagation();
        const name = cell.dataset.emp, idx = parseInt(cell.dataset.idx, 10);
        const options = DATA.weekly_codes.map(v => ({ value: v, label: v, color: DATA.shift_colors[v] }));
        openPicker(cell, options, (val) => {
          DATA.weekly[name][idx] = val;
          renderAll();
        });
      });
    });
  }
}

// --- kivételek ---
function renderExceptions() {
  const search = document.getElementById('searchExc').value.trim().toLowerCase();
  const rows = [];
  DATA.employees.forEach(name => {
    const vals = DATA.weekly[name] || [];
    DATA.week_labels.forEach((label, i) => {
      if (DATA.week_years[i] !== currentYear) return;
      if (search && !name.toLowerCase().includes(search)) return;
      const actual = vals[i] || '';
      const exp = expectedShift(label);
      if (actual && actual !== exp) rows.push({ employee: name, week: label, expected: exp, actual });
    });
  });

  let html = '<thead><tr><th class="name-col">Munkatárs</th><th>Hét</th><th>Elvárt (rotáció)</th><th>Tényleges</th></tr></thead><tbody>';
  if (rows.length === 0) {
    html += '<tr><td colspan="4" style="text-align:center;color:#6b7280;">Nincs kivétel ebben az évben.</td></tr>';
  }
  rows.forEach(e => {
    const parts = e.week.split('.');
    const color = DATA.shift_colors[e.actual];
    const style = color ? `background:${color}${e.actual === 'Éjjel' ? ';color:white' : ''};` : '';
    html += `<tr><td class="name-col">${e.employee}</td><td>${parts[1]}.${parts[2]}</td><td>${e.expected}</td><td style="${style}">${e.actual}</td></tr>`;
  });
  html += '</tbody>';
  document.getElementById('excTable').innerHTML = html;
}

// --- éves összesítő ---
function renderSummary() {
  const container = document.getElementById('summaryCards');
  container.innerHTML = '';
  const labelMap = { SZ: 'Szabadság', B: 'Betegszabadság', OH: 'Home office', UN: 'Ünnepnap', P: 'Pihenőnap' };
  DATA.employees.forEach(name => {
    const counts = computeSummary(name, currentYear);
    const keret = (DATA.quota[name] && DATA.quota[name][currentYear]) || 0;
    const hatra = keret - counts.SZ;
    const card = document.createElement('div');
    card.className = 'summary-card';
    let rowsHtml = '';
    DATA.summary_codes.forEach(code => {
      rowsHtml += `<tr><td>${labelMap[code]}</td><td class="num">${counts[code]}</td></tr>`;
    });
    const keretInput = editMode
      ? `<input type="number" value="${keret}" data-emp="${name}" class="quota-input" style="width:56px;">`
      : `${keret} nap`;
    const nameHtml = editMode
      ? `<span class="editable-name-inline" data-name="${name}" title="Kattints az átnevezéshez"><span class="avatar">${initials(name)}</span>${name} ✎</span>`
      : `<span class="avatar">${initials(name)}</span>${name}`;
    card.innerHTML = `<h3>${nameHtml}</h3><table>${rowsHtml}</table>
      <div class="quota-box">
        <span class="quota-caption">Keret: ${keretInput}</span>
        <span class="quota-remaining">${hatra}<span class="quota-caption"> nap hátra</span></span>
      </div>`;
    container.appendChild(card);
  });

  if (editMode) {
    document.querySelectorAll('.quota-input').forEach(inp => {
      inp.addEventListener('change', () => {
        const name = inp.dataset.emp;
        DATA.quota[name] = DATA.quota[name] || {};
        DATA.quota[name][currentYear] = parseInt(inp.value, 10) || 0;
        renderSummary();
      });
    });
    document.querySelectorAll('.editable-name-inline').forEach(el => {
      el.addEventListener('click', (e) => { e.stopPropagation(); renameEmployee(el.dataset.name); });
    });
  }
}

document.getElementById('minStaffInput').addEventListener('change', (e) => {
  DATA.daily_by_year[currentYear].min_staffing = parseInt(e.target.value, 10) || 0;
  renderDaily();
});

function renderAll() { renderDaily(); renderWeekly(); renderExceptions(); renderSummary(); }

monthSelect.addEventListener('change', renderDaily);
document.getElementById('searchDaily').addEventListener('input', renderDaily);
document.getElementById('searchWeekly').addEventListener('input', renderWeekly);
document.getElementById('searchExc').addEventListener('input', renderExceptions);

renderAll();
</script>
</body>
</html>
"""

# NOTE: admin_html készítése lentebb történik, miután a PERSONAL_TEMPLATE, CSS_BLOCK és
# slugify is definiálva van (az admin oldal saját JS-e beágyazva tartalmazza a
# PERSONAL_TEMPLATE-et, hogy Mentéskor a people/*.html fájlokat is szinkronizálni tudja).

# ---------------------------------------------------------------- publikus főoldal (nincs benne adat)
INDEX_TEMPLATE = """<!DOCTYPE html>
<html lang="hu">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Csapat naptár</title>
<style>
  body { font-family: -apple-system, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    background: linear-gradient(120deg, #4f46e5 0%, #7c3aed 100%); color: white; margin: 0;
    min-height: 100vh; display: flex; align-items: center; justify-content: center; text-align: center; }
  .box { max-width: 420px; padding: 40px; }
  h1 { font-size: 22px; margin-bottom: 10px; }
  p { opacity: 0.85; font-size: 14px; line-height: 1.6; }
</style>
</head>
<body>
  <div class="box">
    <h1>Csapat naptár</h1>
    <p>Ez az oldal nem tartalmaz adatot. A saját naptáradat a tőled kapott személyes linken
    éred el.</p>
  </div>
</body>
</html>
"""
with open(INDEX_OUT, "w", encoding="utf-8") as f:
    f.write(INDEX_TEMPLATE)
print("OK:", INDEX_OUT, "(publikus, adatmentes főoldal)")

# ---------------------------------------------------------------- személyes, szűkített oldalak
def slugify(name):
    s = name.lower()
    repl = {"á": "a", "é": "e", "í": "i", "ó": "o", "ö": "o", "ő": "o", "ú": "u", "ü": "u", "ű": "u"}
    for k, v in repl.items():
        s = s.replace(k, v)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s

CSS_BLOCK = re.search(r"<style>(.*?)</style>", HTML_TEMPLATE, re.S).group(1)

PERSONAL_TEMPLATE = """<!DOCTYPE html>
<html lang="hu">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__P_NAME__ — saját naptár</title>
<style>__P_CSS_BLOCK__</style>
</head>
<body>
<script type="application/json" id="app-data">__P_DATA_JSON__</script>
<header>
  <div class="titles">
    <h1>__P_NAME__ — saját naptár</h1>
    <p>Utolsó frissítés: __P_GENERATED__ &nbsp;·&nbsp; csak a saját adataid</p>
  </div>
  <div class="header-actions">
    <div class="today-badge">Ma: __P_TODAY_DISPLAY__</div>
  </div>
</header>
<div class="wrap">
  <div class="tabs">
    <button class="tab-btn active" data-tab="daily">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><rect x="3" y="5" width="18" height="16" rx="2"/><path d="M3 10h18M8 3v4M16 3v4"/></svg>
      Napi jelenlét
    </button>
    <button class="tab-btn" data-tab="weekly">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><circle cx="12" cy="12" r="9"/><path d="M12 7v5l3 3"/></svg>
      Heti beosztás
    </button>
    <button class="tab-btn" data-tab="exceptions">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.3 3.9 1.8 18a2 2 0 0 0 1.7 3h17a2 2 0 0 0 1.7-3L13.7 3.9a2 2 0 0 0-3.4 0Z"/><path d="M12 9v4M12 17h.01"/></svg>
      Kivételek
    </button>
    <button class="tab-btn" data-tab="summary">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><path d="M3 3v18h18M8 17V10M13 17V6M18 17v-4"/></svg>
      Éves összesítő
    </button>
  </div>

  <div id="daily" class="panel active">
    <div class="controls">
      <label>Hónap</label>
      <select id="monthSelect"></select>
      <button class="btn-icon" onclick="window.print()">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M6 9V2h12v7M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2M6 14h12v8H6z"/></svg>
        Nyomtatás / PDF
      </button>
    </div>
    <div class="table-scroll"><table id="dailyTable"></table></div>
    <div class="legend" id="legend"></div>
  </div>

  <div id="weekly" class="panel">
    <div class="table-scroll"><table id="weeklyTable"></table></div>
    <div class="legend" id="legendWeekly"></div>
  </div>

  <div id="exceptions" class="panel">
    <p class="empty-note">Azok a hetek, ahol a te beosztásod eltér az automatikus Éjjel→Délután→Délelőtt rotációtól.</p>
    <div class="table-scroll"><table id="excTable"></table></div>
  </div>

  <div id="summary" class="panel">
    <div class="summary-cards" id="summaryCards"></div>
  </div>

  <div class="year-bar" id="yearBar"></div>
</div>
<footer>Ez az oldal csak a te adataidat tartalmazza. Kérdés esetén keresd a csoportvezetőt.</footer>

<script>
const DATA = JSON.parse(document.getElementById('app-data').textContent);
DATA.employees = [DATA.employee];
let currentYear = DATA.years.includes(DATA.today_year) ? DATA.today_year : DATA.years[0];

document.querySelectorAll('.tab-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById(btn.dataset.tab).classList.add('active');
  });
});

const yearBar = document.getElementById('yearBar');
DATA.years.forEach(y => {
  const btn = document.createElement('button');
  btn.className = 'year-btn' + (y === currentYear ? ' active' : '');
  btn.textContent = y;
  btn.dataset.year = y;
  btn.addEventListener('click', () => {
    currentYear = y;
    document.querySelectorAll('.year-btn').forEach(b => b.classList.toggle('active', parseInt(b.dataset.year) === y));
    renderAll();
  });
  yearBar.appendChild(btn);
});

const legendEl = document.getElementById('legend');
Object.entries(DATA.codes).forEach(([code, info]) => {
  if (!code) return;
  const el = document.createElement('span');
  el.className = 'chip';
  el.innerHTML = `<span class="swatch" style="background:${info.color}"></span>${info.label} (${code})`;
  legendEl.appendChild(el);
});
const legendWeeklyEl = document.getElementById('legendWeekly');
Object.entries(DATA.shift_colors).forEach(([label, color]) => {
  const el = document.createElement('span');
  el.className = 'chip';
  el.innerHTML = `<span class="swatch" style="background:${color}"></span>${label}`;
  legendWeeklyEl.appendChild(el);
});

const monthNames = ['Január','Február','Március','Április','Május','Június','Július','Augusztus','Szeptember','Október','November','December'];
const huMonthAbbr = ['jan','feb','márc','ápr','máj','jún','júl','aug','szep','okt','nov','dec'];
const monthSelect = document.getElementById('monthSelect');
monthNames.forEach((name, idx) => {
  const opt = document.createElement('option');
  opt.value = idx; opt.textContent = name;
  monthSelect.appendChild(opt);
});
monthSelect.value = huMonthAbbr.indexOf(DATA.today_label.split('.')[0]);

function monthOfIndex(dates, i) { return huMonthAbbr.indexOf(dates[i].split('.')[0]); }

function expectedShift(weekLabel) {
  const [y, mo, da] = weekLabel.split('.').map(Number);
  const [ty, tmo, tda] = DATA.today_week_label.split('.').map(Number);
  const days = Math.round((Date.UTC(y, mo - 1, da) - Date.UTC(ty, tmo - 1, tda)) / 86400000);
  const offset = Math.round(days / 7);
  const idx = ((offset % 3) + 3) % 3;
  return ['Éjjel', 'Délután', 'Délelőtt'][idx];
}
function computeSummary(name, year) {
  const yd = DATA.daily_by_year[year];
  const counts = { SZ: 0, B: 0, OH: 0, UN: 0, P: 0 };
  (yd.daily[name] || []).forEach(c => { if (counts[c] !== undefined) counts[c]++; });
  return counts;
}

function renderDaily() {
  const yearData = DATA.daily_by_year[currentYear];
  if (!yearData) return;
  const monthIdx = parseInt(monthSelect.value, 10);
  const colIndexes = [];
  yearData.dates.forEach((d, i) => { if (monthOfIndex(yearData.dates, i) === monthIdx) colIndexes.push(i); });
  const isToday = (i) => currentYear === DATA.today_year && yearData.dates[i] === DATA.today_label;

  let html = '<thead><tr><th class="name-col">Nap</th>';
  colIndexes.forEach(i => { html += `<th class="${isToday(i) ? 'today-col' : ''}">${yearData.dates[i].split('.')[1]}</th>`; });
  html += '</tr></thead><tbody><tr><td class="name-col">' + DATA.employee + '</td>';
  colIndexes.forEach(i => {
    const code = (yearData.daily[DATA.employee] && yearData.daily[DATA.employee][i]) || '';
    const info = DATA.codes[code];
    const bg = (info && code) ? `background:${info.color}${code==='B'?';color:white':''};` : '';
    html += `<td class="${isToday(i) ? 'today-col' : ''}" style="${bg}" title="${info ? info.label : ''}">${code}</td>`;
  });
  html += '</tr></tbody>';
  document.getElementById('dailyTable').innerHTML = html;
}

function renderWeekly() {
  const colIndexes = [];
  DATA.week_labels.forEach((w, i) => { if (DATA.week_years[i] === currentYear) colIndexes.push(i); });
  let html = '<thead><tr><th class="name-col">Hét</th>';
  colIndexes.forEach(i => {
    const parts = DATA.week_labels[i].split('.');
    const isToday = DATA.week_labels[i] === DATA.today_week_label;
    html += `<th class="${isToday ? 'today-col' : ''}">${parts[1]}.${parts[2]}</th>`;
  });
  html += '</tr></thead><tbody><tr><td class="name-col">' + DATA.employee + '</td>';
  const vals = DATA.weekly[DATA.employee] || [];
  colIndexes.forEach(i => {
    const v = vals[i] || '';
    const color = DATA.shift_colors[v];
    const isToday = DATA.week_labels[i] === DATA.today_week_label;
    const style = color ? `background:${color}${v === 'Éjjel' ? ';color:white' : ''};` : '';
    html += `<td class="${isToday ? 'today-col' : ''}" style="${style}">${v}</td>`;
  });
  html += '</tr></tbody>';
  document.getElementById('weeklyTable').innerHTML = html;
}

function renderExceptions() {
  const rows = [];
  const vals = DATA.weekly[DATA.employee] || [];
  DATA.week_labels.forEach((label, i) => {
    if (DATA.week_years[i] !== currentYear) return;
    const actual = vals[i] || '';
    const exp = expectedShift(label);
    if (actual && actual !== exp) rows.push({ week: label, expected: exp, actual });
  });
  let html = '<thead><tr><th>Hét</th><th>Elvárt (rotáció)</th><th>Tényleges</th></tr></thead><tbody>';
  if (rows.length === 0) html += '<tr><td colspan="3" style="text-align:center;color:#6b7280;">Nincs kivétel ebben az évben.</td></tr>';
  rows.forEach(e => {
    const parts = e.week.split('.');
    const color = DATA.shift_colors[e.actual];
    const style = color ? `background:${color}${e.actual === 'Éjjel' ? ';color:white' : ''};` : '';
    html += `<tr><td>${parts[1]}.${parts[2]}</td><td>${e.expected}</td><td style="${style}">${e.actual}</td></tr>`;
  });
  html += '</tbody>';
  document.getElementById('excTable').innerHTML = html;
}

function renderSummary() {
  const container = document.getElementById('summaryCards');
  const labelMap = { SZ: 'Szabadság', B: 'Betegszabadság', OH: 'Home office', UN: 'Ünnepnap', P: 'Pihenőnap' };
  const counts = computeSummary(DATA.employee, currentYear);
  const keret = (DATA.quota[DATA.employee] && DATA.quota[DATA.employee][currentYear]) || 0;
  const hatra = keret - counts.SZ;
  let rowsHtml = '';
  DATA.summary_codes.forEach(code => { rowsHtml += `<tr><td>${labelMap[code]}</td><td class="num">${counts[code]}</td></tr>`; });
  container.innerHTML = `<div class="summary-card" style="flex-basis:280px;"><h3>${DATA.employee}</h3><table>${rowsHtml}</table>
    <div class="quota-box"><span class="quota-caption">Keret: ${keret} nap</span>
    <span class="quota-remaining">${hatra}<span class="quota-caption"> nap hátra</span></span></div></div>`;
}

function renderAll() { renderDaily(); renderWeekly(); renderExceptions(); renderSummary(); }
monthSelect.addEventListener('change', renderDaily);
renderAll();
</script>
</body>
</html>
"""

# ---------------------------------------------------------------- admin.html végleges legyártása
# (itt már elérhető a PERSONAL_TEMPLATE és a slugify, ezeket beágyazzuk az admin oldal JS-ébe,
# hogy a Mentés gomb a people/*.html fájlokat is tudja frissíteni a GitHub-on)
# FONTOS: a __DATA_JSON__-t (admin saját adata) előbb kell cserélni, mint a
# __PERSONAL_TEMPLATE_JSON__-t, mert a beágyazott PERSONAL_TEMPLATE szövege is tartalmazza
# a "__DATA_JSON__" szót (saját, JS-oldali cseréhez) — ha fordítva csinálnánk, az utólagos
# __DATA_JSON__ csere véletlenül a beágyazott sablonban lévő placeholdert is felülírná.
# A PERSONAL_TEMPLATE saját <script>/</script> tageket tartalmaz (a people/ oldalak
# JS-e). Ha ezt szó szerint (escapelés nélkül) ágyaznánk be egy JSON stringként az
# admin.html <script> blokkjába, a böngésző HTML-parsere a beágyazott "</script>"
# szövegnél lezárná az admin oldal script blokkját (a HTML tokenizer erre a
# karaktersorozatra figyel, függetlenül attól, hogy az JS stringen belül van-e) — ezért
# minden "</" előfordulást "<\\/"-re cserélünk a beágyazott JSON szövegben.
personal_template_json = json.dumps(PERSONAL_TEMPLATE, ensure_ascii=False).replace("</", "<\\/")

admin_html = (HTML_TEMPLATE
        .replace("__GENERATED__", data["generated"])
        .replace("__TODAY_DISPLAY__", data["today_display"])
        .replace("__DATA_JSON__", json.dumps(data, ensure_ascii=False))
        .replace("__PERSONAL_TEMPLATE_JSON__", personal_template_json))
with open(OUT, "w", encoding="utf-8") as f:
    f.write(admin_html)
print("OK:", OUT, "(teljes nézet — csak a megbízott adminoknak küldd el a linkjét, ne linkeld be sehonnan)")

os.makedirs(PEOPLE_DIR, exist_ok=True)
for name in all_employees:
    p_daily_by_year = {
        y: {"dates": daily_by_year[y]["dates"], "daily": {name: daily_by_year[y]["daily"][name]}}
        for y in years
    }
    p_data = {
        "generated": data["generated"],
        "years": years,
        "employee": name,
        "daily_by_year": p_daily_by_year,
        "week_labels": week_labels,
        "week_years": week_years,
        "weekly": {name: weekly.get(name, [])},
        "codes": DAILY_CODES,
        "summary_codes": SUMMARY_CODES,
        "shift_colors": SHIFT_COLORS,
        "quota": {name: quota.get(name, {})},
        "today_year": data["today_year"],
        "today_label": data["today_label"],
        "today_week_label": data["today_week_label"],
        "today_display": data["today_display"],
    }
    p_html = (PERSONAL_TEMPLATE
              .replace("__P_CSS_BLOCK__", CSS_BLOCK)
              .replace("__P_NAME__", name)
              .replace("__P_GENERATED__", p_data["generated"])
              .replace("__P_TODAY_DISPLAY__", p_data["today_display"])
              .replace("__P_DATA_JSON__", json.dumps(p_data, ensure_ascii=False)))
    fname = os.path.join(PEOPLE_DIR, f"{slugify(name)}.html")
    with open(fname, "w", encoding="utf-8") as f:
        f.write(p_html)

print(f"OK: {len(all_employees)} személyes oldal a '{PEOPLE_DIR}/' mappában (mindegyik csak a saját adatát tartalmazza)")
